from fastapi import FastAPI, APIRouter, File, UploadFile, HTTPException
from fastapi.responses import JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Dict, Any, Optional
import uuid
from datetime import datetime, timezone
import pandas as pd
import io
import json

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Data Models
class LiquorData(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    brand_name: str
    rate: float
    daily_sales: Dict[str, int] = Field(default_factory=dict)
    monthly_sale_qty: int
    monthly_sale_value: float
    avg_daily_sale: float
    stock_available_days: float
    stock_value_before: float
    stock_value_today: float
    stock_ratio: float
    # New fields for enhanced analysis
    index_number: int = Field(default=0)
    wholesale_rate: float = Field(default=0.0)
    selling_rate: float = Field(default=0.0)
    D1_date: str = Field(default="N/A")
    D1_stock: float = Field(default=0.0)
    DL_date: str = Field(default="N/A")
    DL_stock: float = Field(default=0.0)
    total_sales_qty: float = Field(default=0.0)
    avg_daily_sales_qty: float = Field(default=0.0)
    days_analyzed: int = Field(default=0)
    current_stock_qty: int = Field(default=0)
    upload_timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class OverstockConfig(BaseModel):
    multiplier: float = 3.0

class AnalyticsResponse(BaseModel):
    total_brands: int
    total_stock_value: float
    total_overstocked_value: float
    overstocked_brands: int
    top_selling_brands: List[Dict[str, Any]]
    overstocked_items: List[Dict[str, Any]]
    sales_trends: Dict[str, Any]

class ChartsResponse(BaseModel):
    volume_leaders: List[Dict[str, Any]]
    velocity_leaders: List[Dict[str, Any]]
    revenue_leaders: List[Dict[str, Any]]
    revenue_proportion: List[Dict[str, Any]]

class DemandRecommendation(BaseModel):
    brand_name: str
    selling_rate: float
    wholesale_rate: float
    current_stock_qty: int
    recommended_qty: int
    urgency_level: str

# Helper functions
def parse_excel_data(file_content: bytes) -> List[Dict[str, Any]]:
    """Parse Excel file and return structured data - supports both tabular and list formats"""
    try:
        # Try to read as Excel with headers first (tabular format)
        try:
            df = pd.read_excel(io.BytesIO(file_content))
            
            # Check if it looks like a tabular format (has typical column names)
            if len(df.columns) >= 3 and any(col.lower().strip() in ['brand name', 'brand_name', 'product', 'name'] for col in df.columns):
                return parse_tabular_format(df)
            else:
                # Try headerless format
                df_headerless = pd.read_excel(io.BytesIO(file_content), header=None)
                return parse_list_format(df_headerless)
                
        except Exception:
            # If Excel fails, try CSV
            try:
                df = pd.read_csv(io.BytesIO(file_content))
                if len(df.columns) >= 3 and any(col.lower().strip() in ['brand name', 'brand_name', 'product', 'name'] for col in df.columns):
                    return parse_tabular_format(df)
                else:
                    df_headerless = pd.read_csv(io.BytesIO(file_content), header=None)
                    return parse_list_format(df_headerless)
            except Exception as csv_error:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Unable to parse file. Please ensure it's a valid Excel or CSV file. Error: {str(csv_error)}"
                )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error parsing file: {str(e)}")

def parse_tabular_format(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """Parse liquor stock data with proper daily stock analysis"""
    if df.empty:
        raise HTTPException(status_code=400, detail="The uploaded file is empty or contains no data")
    
    # Clean column names
    df.columns = df.columns.str.strip()
    
    # Find key columns
    brand_col = None
    wholesale_rate_col = None
    selling_rate_col = None
    index_col = None
    date_columns = []
    
    for col in df.columns:
        col_lower = col.lower().strip()
        if 'brand' in col_lower and 'name' in col_lower:
            brand_col = col
        elif 'wholesale' in col_lower and 'rate' in col_lower:
            wholesale_rate_col = col
        elif ('selling' in col_lower or 'retail' in col_lower) and 'rate' in col_lower:
            selling_rate_col = col
        elif 'rate' in col_lower and not wholesale_rate_col and not selling_rate_col:
            selling_rate_col = col  # Default to selling rate if only one rate column
        elif any(term in col_lower for term in ['index', 'sl', 'sr', 'no', 'id']) and len(col) <= 10:
            index_col = col
        else:
            # Check if column represents a date (contains date patterns)
            if any(date_part in col_lower for date_part in ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']):
                # Sort dates chronologically
                date_columns.append(col)
    
    # Sort date columns chronologically
    date_columns.sort()
    
    print(f"Detected columns - Brand: {brand_col}, Index: {index_col}, Wholesale: {wholesale_rate_col}, Selling: {selling_rate_col}")
    print(f"Date columns found: {date_columns}")
    
    if not brand_col:
        raise HTTPException(status_code=400, detail="Could not find 'Brand Name' column in the file")
    
    if not date_columns:
        raise HTTPException(status_code=400, detail="Could not find date columns for daily stock data")
    
    # Filter out only obvious header rows, be more lenient
    df = df[df[brand_col].notna()]
    df = df[~df[brand_col].astype(str).str.contains('total|sum|^brand name$|^name$|^brand$', na=False, case=False)]
    
    print(f"After filtering, found {len(df)} potential brand rows")  # Debug info
    
    if df.empty:
        raise HTTPException(status_code=400, detail="No valid brand data found after filtering")
    
    # STEP 1: Find global D1 date (when stock increased for ANY brand)
    global_D1_date = None
    print(f"Finding global D1 date across all brands...")
    
    # Collect all stock data across all brands first
    all_brand_stock_data = {}
    for idx, row in df.iterrows():
        brand_name = str(row[brand_col]).strip()
        if not brand_name or brand_name.lower() in ['nan', 'none', '']:
            continue
            
        brand_stock_data = {}
        for date_col in date_columns:
            if pd.notna(row[date_col]):
                try:
                    stock_qty = float(row[date_col])
                    brand_stock_data[date_col] = stock_qty
                except:
                    brand_stock_data[date_col] = 0
        
        if brand_stock_data:
            all_brand_stock_data[brand_name] = brand_stock_data
    
    # Find D1: Look for the date when ANY brand shows stock increase
    sorted_dates = sorted(date_columns)
    for i in range(1, len(sorted_dates)):
        prev_date = sorted_dates[i-1] 
        curr_date = sorted_dates[i]
        
        # Check if any brand shows stock increase on curr_date
        for brand_name, stock_data in all_brand_stock_data.items():
            if prev_date in stock_data and curr_date in stock_data:
                prev_stock = stock_data[prev_date]
                curr_stock = stock_data[curr_date]
                
                # Stock increase indicates new arrival (threshold: any increase)
                if curr_stock > prev_stock:
                    global_D1_date = curr_date
                    print(f"Found global D1: {curr_date} (Brand: {brand_name}, Stock: {prev_stock} -> {curr_stock})")
                    break
        
        if global_D1_date:
            break
    
    # If no stock increase found, use first date
    if not global_D1_date:
        global_D1_date = sorted_dates[0]
        print(f"No stock increase found, using first date as D1: {global_D1_date}")
    
    print(f"Global D1 determined: {global_D1_date}")
    
    # STEP 2: Process each brand with the global D1
    liquor_data = []
    
    for idx, row in df.iterrows():
        try:
            brand_name = str(row[brand_col]).strip()
            if not brand_name or brand_name.lower() in ['nan', 'none', '']:
                continue
            
            # Get index - preserve original index from Excel file
            index_num = idx + 1
            if index_col and pd.notna(row[index_col]):
                try:
                    # Try to extract integer from the index value
                    index_val = str(row[index_col]).strip()
                    # Handle cases where index might have extra characters
                    import re
                    numeric_match = re.search(r'\d+', index_val)
                    if numeric_match:
                        index_num = int(numeric_match.group())
                    else:
                        index_num = int(float(row[index_col]))
                except:
                    index_num = idx + 1
            
            # Get rates
            wholesale_rate = 0.0
            selling_rate = 0.0
            
            if wholesale_rate_col and pd.notna(row[wholesale_rate_col]):
                try:
                    wholesale_rate = float(row[wholesale_rate_col])
                except:
                    pass
            
            if selling_rate_col and pd.notna(row[selling_rate_col]):
                try:
                    selling_rate = float(row[selling_rate_col])
                except:
                    pass
            
            # If wholesale rate not provided, calculate as 90% of selling rate
            if wholesale_rate == 0 and selling_rate > 0:
                wholesale_rate = selling_rate * 0.9
            elif selling_rate == 0 and wholesale_rate > 0:
                selling_rate = wholesale_rate / 0.9
            
            # Get daily stock data with enhanced debugging
            daily_stock_data = {}
            valid_stock_values = []
            
            print(f"Processing {brand_name} - Available date columns: {date_columns}")
            
            for date_col in date_columns:
                try:
                    raw_value = row[date_col]
                    print(f"  {date_col}: raw_value = {raw_value}, type = {type(raw_value)}, notna = {pd.notna(raw_value)}")
                    
                    if pd.notna(raw_value) and str(raw_value).strip() != '':
                        # Try multiple conversion methods
                        stock_qty = 0
                        try:
                            stock_qty = float(raw_value)
                        except (ValueError, TypeError):
                            try:
                                # Handle string numbers
                                stock_qty = float(str(raw_value).replace(',', ''))
                            except:
                                stock_qty = 0
                        
                        daily_stock_data[date_col] = stock_qty
                        print(f"    Converted to: {stock_qty}")
                        
                        if stock_qty >= 0:  # Include zero values too
                            valid_stock_values.append((date_col, stock_qty))
                    else:
                        daily_stock_data[date_col] = 0
                        print(f"    Skipped (empty/NaN)")
                        
                except Exception as e:
                    daily_stock_data[date_col] = 0
                    print(f"    Error processing {date_col}: {e}")
            
            print(f"  Final valid_stock_values: {valid_stock_values}")
            
            if not valid_stock_values:
                print(f"  WARNING: No valid stock data found for {brand_name}")
                continue  # Skip if no valid stock data
            
            # Use GLOBAL D1 and individual DL
            # D1: Use the global D1 date we found earlier
            # DL: Last date with stock data for this specific brand
            
            # Sort by date order (chronologically)
            sorted_stock_values = sorted(valid_stock_values, key=lambda x: x[0])
            
            # Find D1 stock for this brand on the global D1 date
            D1_date = global_D1_date
            D1_stock = daily_stock_data.get(global_D1_date, 0)
            
            # If brand has no data on global D1 date, find closest previous date
            if D1_stock == 0:
                for date_col, stock_val in sorted_stock_values:
                    if date_col <= global_D1_date:
                        D1_stock = stock_val
                    else:
                        break
            
            # DL: Last date with stock data for this brand
            DL_date, DL_stock = sorted_stock_values[-1]
            
            print(f"  Using Global D1: {D1_date} = {D1_stock}, Individual DL: {DL_date} = {DL_stock}")
            
            # Calculate total sales between D1 and DL
            total_sales_qty = max(0, D1_stock - DL_stock)  # Stock reduction = sales
            
            # Calculate number of days between D1 and DL
            days_between = max(1, len(sorted_stock_values) - 1)
            
            # Calculate average daily sales
            avg_daily_sales_qty = total_sales_qty / days_between if days_between > 0 else 0
            
            # Calculate monthly sales (24 days as requested)
            monthly_sales_qty = avg_daily_sales_qty * 24
            monthly_sales_value = monthly_sales_qty * selling_rate
            
            # Current stock value (on DL date)
            current_stock_value = DL_stock * selling_rate
            
            # Calculate stock ratio (stock value / monthly sales value)
            stock_ratio = current_stock_value / max(1, monthly_sales_value) if monthly_sales_value > 0 else 0
            
            # Stock availability in days
            stock_available_days = (DL_stock / max(0.1, avg_daily_sales_qty)) if avg_daily_sales_qty > 0 else 999
            
            print(f"  Calculations: Total sales={total_sales_qty}, Days={days_between}, Avg daily={avg_daily_sales_qty:.2f}")
            print(f"  Monthly sales value={monthly_sales_value:.2f}, Stock ratio={stock_ratio:.2f}")
            
            brand_data = {
                'brand_name': brand_name,
                'product_id': f"ID_{index_num}",
                'index_number': int(index_num),
                'wholesale_rate': float(wholesale_rate),
                'selling_rate': float(selling_rate),
                'rate': float(selling_rate),  # For compatibility
                'D1_date': str(D1_date),
                'D1_stock': float(D1_stock),
                'DL_date': str(DL_date),
                'DL_stock': float(DL_stock),
                'current_stock_qty': int(max(0, DL_stock)),
                'total_sales_qty': float(total_sales_qty),
                'avg_daily_sales_qty': float(avg_daily_sales_qty),
                'monthly_sales_qty': float(monthly_sales_qty),
                'monthly_sale_value': float(monthly_sales_value),
                'monthly_sale_qty': int(max(0, monthly_sales_qty)),  # For compatibility
                'stock_value_today': float(current_stock_value),
                'stock_ratio': float(stock_ratio),
                'stock_available_days': float(min(999, max(0, stock_available_days))),
                'avg_daily_sale': float(monthly_sales_value / 30),  # For compatibility
                'stock_value_before': float(D1_stock * selling_rate),
                'daily_sales': daily_stock_data,
                'days_analyzed': int(max(1, days_between)),
            }
            
            liquor_data.append(brand_data)
            
        except Exception as e:
            logging.error(f"Error parsing row {idx} ({brand_name}): {e}")
            print(f"Failed to parse brand: {brand_name}, Error: {e}")  # Debug info
            continue
    
    if not liquor_data:
        raise HTTPException(status_code=400, detail="No valid liquor data could be extracted from the file")
    
    print(f"Successfully parsed {len(liquor_data)} out of {len(df)} potential brands")
    logging.info(f"Successfully parsed {len(liquor_data)} liquor brands with proper stock analysis")
    return liquor_data

def parse_list_format(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """Parse simple list format (brand names followed by numerical data)"""
    if df.empty:
        raise HTTPException(status_code=400, detail="The uploaded file is empty or contains no data")
    
    # Flatten all data and separate text from numbers
    all_data = df.values.flatten()
    all_data = [str(x).strip() for x in all_data if pd.notna(x) and str(x).strip()]
    
    brand_names = []
    numerical_data = []
    
    for item in all_data:
        try:
            float(item)
            numerical_data.append(item)
        except (ValueError, TypeError):
            if item and not item.replace('.', '').isdigit():
                brand_names.append(item)
    
    if not brand_names:
        raise HTTPException(status_code=400, detail="No brand names found in the file")
    
    if len(numerical_data) < len(brand_names) * 2:
        raise HTTPException(
            status_code=400, 
            detail=f"Insufficient numerical data. Expected at least {len(brand_names) * 2} values, found {len(numerical_data)}"
        )
    
    liquor_data = []
    
    # Parse in groups of 3 (ID, Rate, Quantity)
    for i in range(len(brand_names)):
        try:
            if i * 3 < len(numerical_data):
                product_id = numerical_data[i * 3] if i * 3 < len(numerical_data) else f"AUTO_{i}"
                rate = float(numerical_data[i * 3 + 1]) if i * 3 + 1 < len(numerical_data) else 100.0
                quantity = int(float(numerical_data[i * 3 + 2])) if i * 3 + 2 < len(numerical_data) else 0
                
                estimated_monthly_sales = max(1, quantity // 4) * rate
                
                brand_data = {
                    'brand_name': brand_names[i],
                    'product_id': product_id,
                    'rate': rate,
                    'current_stock_qty': quantity,
                    'stock_value_today': rate * quantity,
                    'monthly_sale_value': estimated_monthly_sales,
                    'stock_available_days': max(1, quantity // max(1, quantity // 30)) if quantity > 0 else 0,
                    'stock_ratio': (rate * quantity) / max(1, estimated_monthly_sales),
                    'daily_sales': {},
                    'monthly_sale_qty': max(1, quantity // 4),
                    'avg_daily_sale': estimated_monthly_sales / 30,
                    'stock_value_before': rate * quantity * 1.1,
                }
                
                liquor_data.append(brand_data)
                
        except Exception as e:
            logging.warning(f"Error parsing data for {brand_names[i] if i < len(brand_names) else 'Unknown'}: {e}")
            continue
    
    return liquor_data

def calculate_overstocking(data: List[Dict], multiplier: float = 3.0) -> List[Dict]:
    """Calculate overstocking based on configurable multiplier"""
    overstocked_items = []
    
    for item in data:
        monthly_avg_sale = item.get('monthly_sale_value', 0)
        current_stock_value = item.get('stock_value_today', 0)
        
        # Calculate threshold (multiplier * monthly average)
        threshold = monthly_avg_sale * multiplier
        
        if current_stock_value > threshold and monthly_avg_sale > 0:
            overstock_value = current_stock_value - threshold
            overstocked_items.append({
                'brand_name': item['brand_name'],
                'current_stock_value': current_stock_value,
                'monthly_avg_sale': monthly_avg_sale,
                'threshold': threshold,
                'overstock_value': overstock_value,
                'stock_ratio': item.get('stock_ratio', 0)
            })
    
    return sorted(overstocked_items, key=lambda x: x['overstock_value'], reverse=True)

# API Endpoints
@api_router.get("/")
async def root():
    return {"message": "Liquor Sales Analysis Dashboard API"}

@api_router.post("/upload-data")
async def upload_liquor_data(file: UploadFile = File(...)):
    """Upload and process Excel/CSV file with liquor data"""
    try:
        # Validate file type first
        if not file.filename or not file.filename.endswith(('.xlsx', '.xls', '.csv')):
            raise HTTPException(
                status_code=400, 
                detail={
                    "error": "Invalid file type",
                    "message": "Only Excel (.xlsx, .xls) and CSV files are supported",
                    "supported_formats": [".xlsx", ".xls", ".csv"]
                }
            )
        
        # Read file content
        content = await file.read()
        
        if len(content) == 0:
            raise HTTPException(status_code=400, detail="Empty file uploaded")
        
        # Parse the data
        parsed_data = parse_excel_data(content)
        
        if not parsed_data:
            raise HTTPException(status_code=400, detail="No valid data found in the file")
        
        # Clear existing data and insert new data
        await db.liquor_data.delete_many({})
        
        # Convert to LiquorData models and insert
        liquor_objects = []
        for item_data in parsed_data:
            liquor_obj = LiquorData(**item_data)
            liquor_objects.append(liquor_obj.dict())
        
        if liquor_objects:
            await db.liquor_data.insert_many(liquor_objects)
        
        return JSONResponse(
            status_code=200,
            content={
                "message": f"Successfully uploaded {len(liquor_objects)} liquor records",
                "total_records": len(liquor_objects)
            }
        )
        
    except HTTPException:
        # Re-raise HTTPExceptions (400 errors) as-is
        raise
    except Exception as e:
        logging.error(f"Error uploading data: {e}")
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@api_router.get("/analytics", response_model=AnalyticsResponse)
async def get_analytics(overstock_multiplier: float = 3.0):
    """Get comprehensive analytics including overstocking analysis"""
    try:
        # Fetch all liquor data
        liquor_records = await db.liquor_data.find().to_list(1000)
        
        if not liquor_records:
            raise HTTPException(status_code=404, detail="No data found. Please upload liquor data first.")
        
        # Convert to dict format for calculations
        data_dicts = [
            {
                'brand_name': record['brand_name'],
                'rate': record['rate'],
                'daily_sales': record['daily_sales'],
                'monthly_sale_qty': record['monthly_sale_qty'],
                'monthly_sale_value': record['monthly_sale_value'],
                'avg_daily_sale': record['avg_daily_sale'],
                'stock_available_days': record['stock_available_days'],
                'stock_value_before': record['stock_value_before'],
                'stock_value_today': record['stock_value_today'],
                'stock_ratio': record['stock_ratio']
            }
            for record in liquor_records
        ]
        
        # Calculate overstocked items
        overstocked_items = calculate_overstocking(data_dicts, overstock_multiplier)
        
        # Calculate top selling brands
        top_selling = sorted(data_dicts, key=lambda x: x['monthly_sale_value'], reverse=True)[:10]
        
        # Calculate totals
        total_stock_value = sum(item['stock_value_today'] for item in data_dicts)
        total_overstocked_value = sum(item['overstock_value'] for item in overstocked_items)
        
        # Prepare sales trends data
        sales_trends = {}
        for record in data_dicts:
            for date, sales in record['daily_sales'].items():
                if date not in sales_trends:
                    sales_trends[date] = 0
                sales_trends[date] += sales
        
        # Sort sales trends by date
        sorted_trends = dict(sorted(sales_trends.items()))
        
        return AnalyticsResponse(
            total_brands=len(data_dicts),
            total_stock_value=total_stock_value,
            total_overstocked_value=total_overstocked_value,
            overstocked_brands=len(overstocked_items),
            top_selling_brands=[
                {
                    'brand_name': item['brand_name'],
                    'monthly_sale_value': item['monthly_sale_value'],
                    'stock_value_today': item['stock_value_today'],
                    'stock_ratio': item['stock_ratio']
                }
                for item in top_selling
            ],
            overstocked_items=overstocked_items,
            sales_trends=sorted_trends
        )
        
    except Exception as e:
        logging.error(f"Error getting analytics: {e}")
        raise HTTPException(status_code=500, detail=f"Error calculating analytics: {str(e)}")

@api_router.get("/brands")
async def get_all_brands():
    """Get all brand data"""
    try:
        liquor_records = await db.liquor_data.find().to_list(1000)
        return [LiquorData(**record) for record in liquor_records]
    except Exception as e:
        logging.error(f"Error fetching brands: {e}")
        raise HTTPException(status_code=500, detail=f"Error fetching brands: {str(e)}")

@api_router.get("/charts", response_model=ChartsResponse)
async def get_charts_data():
    """Get data for performance charts and visualizations"""
    try:
        liquor_records = await db.liquor_data.find().to_list(1000)
        
        if not liquor_records:
            raise HTTPException(status_code=404, detail="No data found")
        
        # Convert to dict format for calculations
        data_dicts = [
            {
                'brand_name': record['brand_name'],
                'rate': record['rate'],
                'current_stock_qty': record.get('current_stock_qty', 0),
                'monthly_sale_value': record['monthly_sale_value'],
                'stock_value_today': record['stock_value_today'],
                'stock_available_days': record['stock_available_days'],
                'stock_ratio': record['stock_ratio']
            }
            for record in liquor_records
        ]
        
        # Calculate total sales for proportion
        total_sales = sum(item['monthly_sale_value'] for item in data_dicts)
        
        # Volume Leaders (by current stock quantity)
        volume_leaders = sorted(data_dicts, key=lambda x: x.get('current_stock_qty', 0), reverse=True)[:10]
        volume_chart = [
            {
                'name': item['brand_name'][:20],  # Truncate long names
                'value': item.get('current_stock_qty', 0),
                'stock_value': item['stock_value_today']
            }
            for item in volume_leaders if item.get('current_stock_qty', 0) > 0
        ]
        
        # Velocity Leaders (by stock turnover rate)
        velocity_leaders = [item for item in data_dicts if item['stock_available_days'] > 0]
        velocity_leaders = sorted(velocity_leaders, key=lambda x: x['stock_available_days'])[:10]
        velocity_chart = [
            {
                'name': item['brand_name'],
                'velocity': round(30 / item['stock_available_days'], 2) if item['stock_available_days'] > 0 else 0,
                'days_of_stock': item['stock_available_days'],
                'sales_value': item['monthly_sale_value']
            }
            for item in velocity_leaders
        ]
        
        # Revenue Leaders (by estimated sales value)
        revenue_leaders = sorted(data_dicts, key=lambda x: x['monthly_sale_value'], reverse=True)[:10]
        revenue_chart = [
            {
                'name': item['brand_name'],
                'value': item['monthly_sale_value'],
                'stock_value': item['stock_value_today'],
                'stock_ratio': item['stock_ratio']
            }
            for item in revenue_leaders
        ]
        
        # Revenue Proportion (percentage of total estimated sales)
        revenue_proportion = [
            {
                'name': item['brand_name'],
                'value': item['monthly_sale_value'],
                'percentage': round((item['monthly_sale_value'] / total_sales) * 100, 2) if total_sales > 0 else 0,
                'stock_value': item['stock_value_today']
            }
            for item in revenue_leaders if item['monthly_sale_value'] > 0
        ]
        
        return ChartsResponse(
            volume_leaders=volume_chart,
            velocity_leaders=velocity_chart,
            revenue_leaders=revenue_chart,
            revenue_proportion=revenue_proportion
        )
        
    except Exception as e:
        logging.error(f"Error getting charts data: {e}")
        raise HTTPException(status_code=500, detail=f"Error getting charts data: {str(e)}")

@api_router.get("/demand-recommendations")
async def get_demand_recommendations():
    """Get smart demand recommendations with wholesale rates and quantities"""
    try:
        liquor_records = await db.liquor_data.find().to_list(1000)
        
        if not liquor_records:
            raise HTTPException(status_code=404, detail="No data found")
        
        recommendations = []
        
        for record in liquor_records:
            brand_name = record['brand_name']
            selling_rate = record.get('selling_rate', record['rate'])
            wholesale_rate = record.get('wholesale_rate', selling_rate * 0.9)
            current_stock_qty = record.get('current_stock_qty', 0)
            avg_daily_sales_qty = record.get('avg_daily_sales_qty', 0)
            stock_days = record['stock_available_days']
            
            # Calculate recommended quantity based on actual sales data
            if avg_daily_sales_qty > 0:
                # Target: maintain 30-45 days of stock based on actual daily sales
                target_days = 30  # Target stock for 30 days
                target_qty = int(avg_daily_sales_qty * target_days)
                
                # Calculate recommended order quantity
                recommended_qty = max(0, target_qty - current_stock_qty)
                
                # Determine urgency level based on days of stock remaining
                if stock_days < 10:
                    urgency = "HIGH"
                elif stock_days < 20:
                    urgency = "MEDIUM"
                elif stock_days < 30:
                    urgency = "LOW"
                else:
                    urgency = "NONE"
                
                # Only include items that need restocking
                if urgency != "NONE" or recommended_qty > 5:  # Only recommend if qty > 5
                    recommendations.append(DemandRecommendation(
                        brand_name=brand_name,
                        selling_rate=selling_rate,
                        wholesale_rate=round(wholesale_rate, 2),
                        current_stock_qty=current_stock_qty,
                        recommended_qty=min(recommended_qty, current_stock_qty * 2),  # Cap at 2x current stock
                        urgency_level=urgency
                    ))
        
        # Sort by urgency (HIGH -> MEDIUM -> LOW) and then by recommended quantity
        urgency_order = {"HIGH": 1, "MEDIUM": 2, "LOW": 3}
        recommendations.sort(key=lambda x: (urgency_order.get(x.urgency_level, 4), -x.recommended_qty))
        
        return recommendations
        
    except Exception as e:
        logging.error(f"Error generating demand recommendations: {e}")
        raise HTTPException(status_code=500, detail=f"Error generating recommendations: {str(e)}")

@api_router.get("/calculation-details")
async def get_calculation_details():
    """Get detailed calculations for all brands for verification"""
    try:
        liquor_records = await db.liquor_data.find().to_list(1000)
        
        if not liquor_records:
            raise HTTPException(status_code=404, detail="No data found")
        
        calculation_details = []
        
        for record in liquor_records:
            # Calculate multiplier value (current stock value / monthly sales value)
            current_stock_value = record.get('stock_value_today', 0)
            monthly_sales_value = record.get('monthly_sale_value', 0)
            multiplier_value = current_stock_value / max(1, monthly_sales_value) if monthly_sales_value > 0 else 0
            
            detail = {
                'index': record.get('index_number', record.get('product_id', 'N/A')),
                'brand_name': record['brand_name'],
                'calculated_wholesale_rate': record.get('wholesale_rate', 0),
                'selling_rate': record.get('selling_rate', record.get('rate', 0)),
                'calculated_avg_monthly_sale': record.get('monthly_sale_value', 0),
                'calculated_current_stock_value': current_stock_value,
                'calculated_multiplier_value': round(multiplier_value, 3),
                # Additional useful fields for verification - use actual stored values
                'D1_date': record.get('D1_date', 'N/A'),
                'D1_stock': record.get('D1_stock', 0),
                'DL_date': record.get('DL_date', 'N/A'), 
                'DL_stock': record.get('DL_stock', 0),
                'total_sales_qty': record.get('total_sales_qty', 0),
                'avg_daily_sales_qty': record.get('avg_daily_sales_qty', 0),
                'days_analyzed': record.get('days_analyzed', 0),
                'stock_available_days': record.get('stock_available_days', 0)
            }
            
            calculation_details.append(detail)
        
        # Sort by index number
        # Sort by index number, handling various formats
        def sort_key(x):
            idx_str = str(x['index'])
            try:
                # Try to extract number from index
                import re
                match = re.search(r'\d+', idx_str)
                return int(match.group()) if match else 999
            except:
                return 999
        
        calculation_details.sort(key=sort_key)
        
        return calculation_details
        
    except Exception as e:
        logging.error(f"Error getting calculation details: {e}")
        raise HTTPException(status_code=500, detail=f"Error getting calculation details: {str(e)}")

@api_router.get("/export-demand-list")
async def export_demand_list():
    """Export demand recommendations with updated format: Index, Brand Name, Wholesale Rate, Quantity in Stock, Quantity to be Demanded"""
    try:
        # Get recommendations
        recommendations_data = await get_demand_recommendations()
        
        if not recommendations_data:
            raise HTTPException(status_code=404, detail="No recommendations to export")
        
        # Convert to updated DataFrame format
        df_data = []
        for index, rec in enumerate(recommendations_data, 1):
            df_data.append({
                'Index': index,
                'Brand Name': rec.brand_name,
                'Wholesale Rate': rec.wholesale_rate,
                'Quantity held in Stock': rec.current_stock_qty,
                'Quantity to be Demanded': rec.recommended_qty
            })
        
        df = pd.DataFrame(df_data)
        
        # Create Excel file in memory
        from io import BytesIO
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Demand Forecast', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Demand Forecast']
            
            # Auto-adjust column widths
            column_widths = {'A': 10, 'B': 40, 'C': 18, 'D': 22, 'E': 25}
            for column_letter, width in column_widths.items():
                worksheet.column_dimensions[column_letter].width = width
            
            # Style headers
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col in range(1, 6):  # 5 columns
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
        
        output.seek(0)
        
        from fastapi.responses import Response
        
        return Response(
            content=output.getvalue(),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=liquor_demand_forecast_{datetime.now().strftime("%Y%m%d")}.xlsx'
            }
        )
        
    except Exception as e:
        logging.error(f"Error exporting demand list: {e}")
        raise HTTPException(status_code=500, detail=f"Error exporting demand list: {str(e)}")

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()